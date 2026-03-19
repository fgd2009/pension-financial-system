# -*- coding: utf-8 -*-
"""
养老金融通报数据自动化处理系统
数据处理核心模块

作者: Matrix Agent
创建日期: 2026-03-19
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from .config import BRANCHES, BRANCH_COUNT, SINGLE_KEY_COLS, COLLECTIVE_KEY_COLS


class DataProcessor:
    """数据处理引擎"""

    def __init__(self):
        self.single_plan_data = None
        self.collective_plan_data = None
        self.pivot_scale = None
        self.pivot_customer = None
        self.kpi_scale = None
        self.kpi_customer = None
        self.base_data = None
        self.aux_calc = None

    # =========================================================================
    # 数据导入模块
    # =========================================================================

    def import_single_plan(self, filepath: str) -> Tuple[bool, str]:
        """导入单一计划明细数据"""
        try:
            df = pd.read_excel(filepath, header=None)
            self.single_plan_data = df
            return True, f"成功导入 {len(df)} 行数据"
        except Exception as e:
            return False, f"导入失败: {str(e)}"

    def import_collective_plan(self, filepath: str) -> Tuple[bool, str]:
        """导入集合计划明细数据"""
        try:
            df = pd.read_excel(filepath, header=None)
            self.collective_plan_data = df
            return True, f"成功导入 {len(df)} 行数据"
        except Exception as e:
            return False, f"导入失败: {str(e)}"

    def import_from_template(self, template_path: str) -> Tuple[bool, str]:
        """从模板导入现有数据"""
        try:
            wb = load_workbook(template_path, data_only=True)

            # 读取透视表数据
            ws_scale = wb['透视-规模汇总']
            ws_customer = wb['透视-客户数汇总']

            self.pivot_scale = self._read_pivot_scale(ws_scale)
            self.pivot_customer = self._read_pivot_customer(ws_customer)

            # 读取KPI表数据
            self.kpi_scale = self._read_kpi_scale(wb['kpi-规模'])
            self.kpi_customer = self._read_kpi_customer(wb['kpi-客户数'])

            return True, "模板数据导入成功"

        except Exception as e:
            return False, f"导入失败: {str(e)}"

    def _read_pivot_scale(self, ws) -> pd.DataFrame:
        """读取透视-规模汇总表"""
        data = []
        for row in range(2, 42):  # 第2行到第41行
            branch = ws.cell(row, 1).value
            if branch and branch in BRANCHES:
                row_data = {
                    '机构名称': branch,
                    '单一受托资产': ws.cell(row, 2).value or 0,
                    '单一投资资产': ws.cell(row, 3).value or 0,
                    '单一个人账户数': ws.cell(row, 4).value or 0,
                    '集合受托资产': ws.cell(row, 5).value or 0,
                    '集合投资资产': ws.cell(row, 6).value or 0,
                    '集合个人账户数': ws.cell(row, 7).value or 0,
                    '受托资产合计': ws.cell(row, 8).value or 0,
                    '投资资产合计': ws.cell(row, 9).value or 0,
                    '账管规模合计': ws.cell(row, 10).value or 0,
                }
                data.append(row_data)
        return pd.DataFrame(data)

    def _read_pivot_customer(self, ws) -> pd.DataFrame:
        """读取透视-客户数汇总表"""
        data = []
        for row in range(2, 42):
            branch = ws.cell(row, 1).value
            if branch and branch in BRANCHES:
                row_data = {
                    '机构名称': branch,
                    '单一计划客户数': ws.cell(row, 2).value or 0,
                    '集合计划客户数': ws.cell(row, 3).value or 0,
                    '客户数合计': ws.cell(row, 4).value or 0,
                }
                data.append(row_data)
        return pd.DataFrame(data)

    def _read_kpi_scale(self, ws) -> pd.DataFrame:
        """读取kpi-规模表"""
        data = []
        for row in range(3, 42):  # 第3行开始
            branch = ws.cell(row, 1).value
            if branch and branch in BRANCHES:
                row_data = {'机构名称': branch}

                # 11月基准 (B-G)
                row_data['nov_trustee'] = ws.cell(row, 2).value or 0
                row_data['nov_account'] = ws.cell(row, 3).value or 0
                row_data['nov_invest'] = ws.cell(row, 4).value or 0
                row_data['nov托管'] = ws.cell(row, 5).value or 0

                # 上月 (H-M)
                row_data['last_trustee'] = ws.cell(row, 8).value or 0
                row_data['last_account'] = ws.cell(row, 9).value or 0
                row_data['last_invest'] = ws.cell(row, 10).value or 0
                row_data['last托管'] = ws.cell(row, 11).value or 0

                # 当月 (N-S)
                row_data['curr_trustee'] = ws.cell(row, 14).value or 0
                row_data['curr_account'] = ws.cell(row, 15).value or 0
                row_data['curr_invest'] = ws.cell(row, 16).value or 0
                row_data['curr托管'] = ws.cell(row, 17).value or 0

                data.append(row_data)
        return pd.DataFrame(data) if data else None

    def _read_kpi_customer(self, ws) -> pd.DataFrame:
        """读取kpi-客户数表"""
        data = []
        for row in range(3, 42):
            branch = ws.cell(row, 1).value
            if branch and branch in BRANCHES:
                row_data = {
                    '机构名称': branch,
                    'nov安心健养': ws.cell(row, 2).value or 0,
                    'nov托管': ws.cell(row, 3).value or 0,
                    'last安心健养': ws.cell(row, 5).value or 0,
                    'last托管': ws.cell(row, 6).value or 0,
                    'curr安心健养': ws.cell(row, 8).value or 0,
                    'curr托管': ws.cell(row, 9).value or 0,
                }
                data.append(row_data)
        return pd.DataFrame(data) if data else None

    # =========================================================================
    # 透视汇总计算
    # =========================================================================

    def calculate_pivot_scale(self) -> pd.DataFrame:
        """计算透视-规模汇总表"""
        if self.single_plan_data is None and self.collective_plan_data is None:
            return None

        result = []
        for branch in BRANCHES:
            row = {'机构名称': branch}

            # 单一计划汇总
            single = self.single_plan_data
            if single is not None:
                mask = single[SINGLE_KEY_COLS['branch']].astype(str) == branch
                row['单一受托资产'] = single.loc[mask, SINGLE_KEY_COLS['trustee_asset']].sum()
                row['单一投资资产'] = single.loc[mask, SINGLE_KEY_COLS['investment_asset']].sum()
                row['单一个人账户数'] = single.loc[mask, SINGLE_KEY_COLS['account_count']].sum()
            else:
                row['单一受托资产'] = 0
                row['单一投资资产'] = 0
                row['单一个人账户数'] = 0

            # 集合计划汇总
            collective = self.collective_plan_data
            if collective is not None:
                mask = collective[COLLECTIVE_KEY_COLS['branch']].astype(str) == branch
                row['集合受托资产'] = collective.loc[mask, COLLECTIVE_KEY_COLS['trustee_asset']].sum()
                row['集合投资资产'] = collective.loc[mask, COLLECTIVE_KEY_COLS['investment_asset']].sum()
                row['集合个人账户数'] = collective.loc[mask, COLLECTIVE_KEY_COLS['account_count']].sum()
            else:
                row['集合受托资产'] = 0
                row['集合投资资产'] = 0
                row['集合个人账户数'] = 0

            # 合计
            row['受托资产合计'] = row['单一受托资产'] + row['集合受托资产']
            row['投资资产合计'] = row['单一投资资产'] + row['集合投资资产']
            row['账管规模合计'] = row['单一个人账户数'] + row['集合个人账户数']

            result.append(row)

        self.pivot_scale = pd.DataFrame(result)
        return self.pivot_scale

    def calculate_pivot_customer(self) -> pd.DataFrame:
        """计算透视-客户数汇总表"""
        if self.single_plan_data is None and self.collective_plan_data is None:
            return None

        result = []
        for branch in BRANCHES:
            row = {'机构名称': branch}

            # 单一计划客户数
            if self.single_plan_data is not None:
                mask = self.single_plan_data[SINGLE_KEY_COLS['branch']].astype(str) == branch
                row['单一计划客户数'] = mask.sum()
            else:
                row['单一计划客户数'] = 0

            # 集合计划客户数
            if self.collective_plan_data is not None:
                mask = self.collective_plan_data[COLLECTIVE_KEY_COLS['branch']].astype(str) == branch
                row['集合计划客户数'] = mask.sum()
            else:
                row['集合计划客户数'] = 0

            row['客户数合计'] = row['单一计划客户数'] + row['集合计划客户数']
            result.append(row)

        self.pivot_customer = pd.DataFrame(result)
        return self.pivot_customer

    # =========================================================================
    # KPI计算
    # =========================================================================

    def calculate_kpi_scale(self) -> pd.DataFrame:
        """计算kpi-规模表"""
        if self.pivot_scale is None:
            return None

        result = []
        for _, pivot_row in self.pivot_scale.iterrows():
            branch = pivot_row['机构名称']
            row = {'机构名称': branch}

            # 当月数据(从透视表获取并转换单位)
            curr_trustee = pivot_row['受托资产合计'] / 100000000  # 元→亿元
            curr_account = pivot_row['账管规模合计'] / 10000       # 户→万户
            curr_invest = pivot_row['投资资产合计'] / 100000000    # 元→亿元

            row['curr_trustee'] = curr_trustee
            row['curr_account'] = curr_account
            row['curr_invest'] = curr_invest
            row['curr托管'] = 0  # 手工填写

            # 计算合计
            row['curr_total_1'] = curr_trustee + curr_account + curr_invest + row['curr托管']
            row['curr_total'] = row['curr托管'] + row['curr_total_1']

            # 增量计算(需要有上月数据才能计算)
            if self.kpi_scale is not None:
                last_row = self.kpi_scale[self.kpi_scale['机构名称'] == branch]
                if not last_row.empty:
                    last_total = last_row['nov_total'].values[0] if 'nov_total' in last_row.columns else 0
                    row['increment'] = row['curr_total'] - last_total

            result.append(row)

        self.kpi_scale = pd.DataFrame(result)
        return self.kpi_scale

    def calculate_kpi_customer(self) -> pd.DataFrame:
        """计算kpi-客户数表"""
        if self.pivot_customer is None:
            return None

        result = []
        for _, pivot_row in self.pivot_customer.iterrows():
            branch = pivot_row['机构名称']
            row = {
                '机构名称': branch,
                'curr安心健养': pivot_row['客户数合计'],  # 从透视表获取
                'curr托管': 0,  # 手工填写
            }
            row['curr_total'] = row['curr安心健养'] + row['curr托管']
            result.append(row)

        self.kpi_customer = pd.DataFrame(result)
        return self.kpi_customer

    # =========================================================================
    # 排名计算
    # =========================================================================

    def calculate_rankings(self, df: pd.DataFrame, col: str, rank_col: str) -> pd.DataFrame:
        """计算排名(含并列处理)"""
        df = df.copy()
        df[rank_col] = df[col].rank(method='min', ascending=False).astype(int)
        return df

    # =========================================================================
    # 数据导出
    # =========================================================================

    def get_summary(self) -> Dict:
        """获取数据汇总摘要"""
        summary = {
            '单一计划行数': len(self.single_plan_data) if self.single_plan_data is not None else 0,
            '集合计划行数': len(self.collective_plan_data) if self.collective_plan_data is not None else 0,
            '透视表已计算': self.pivot_scale is not None,
            '客户数表已计算': self.pivot_customer is not None,
        }
        return summary


# 全局实例
processor = DataProcessor()
