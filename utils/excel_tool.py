# -*- coding: utf-8 -*-
"""
养老金融通报数据自动化处理系统
Excel工具模块

作者: Matrix Agent
创建日期: 2026-03-19
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Tuple
import copy

from core.config import BRANCHES


class ExcelTool:
    """Excel文件处理工具"""

    def __init__(self):
        self.workbook = None
        self.template_path = None

    def load_template(self, template_path: str) -> bool:
        """加载Excel模板"""
        try:
            self.workbook = load_workbook(template_path)
            self.template_path = template_path
            return True
        except Exception as e:
            print(f"加载模板失败: {e}")
            return False

    def save_as(self, output_path: str) -> Tuple[bool, str]:
        """保存为新文件"""
        try:
            self.workbook.save(output_path)
            return True, f"保存成功: {output_path}"
        except Exception as e:
            return False, f"保存失败: {e}"

    # =========================================================================
    # 透视表写入
    # =========================================================================

    def write_pivot_scale(self, data: pd.DataFrame):
        """写入透视-规模汇总表"""
        ws = self.workbook['透视-规模汇总']

        for idx, row in data.iterrows():
            row_num = idx + 2  # 从第2行开始
            ws.cell(row_num, 1, row['机构名称'])
            ws.cell(row_num, 2, row['单一受托资产'])
            ws.cell(row_num, 3, row['单一投资资产'])
            ws.cell(row_num, 4, row['单一个人账户数'])
            ws.cell(row_num, 5, row['集合受托资产'])
            ws.cell(row_num, 6, row['集合投资资产'])
            ws.cell(row_num, 7, row['集合个人账户数'])
            ws.cell(row_num, 8, row['受托资产合计'])
            ws.cell(row_num, 9, row['投资资产合计'])
            ws.cell(row_num, 10, row['账管规模合计'])

    def write_pivot_customer(self, data: pd.DataFrame):
        """写入透视-客户数汇总表"""
        ws = self.workbook['透视-客户数汇总']

        for idx, row in data.iterrows():
            row_num = idx + 2
            ws.cell(row_num, 1, row['机构名称'])
            ws.cell(row_num, 2, row['单一计划客户数'])
            ws.cell(row_num, 3, row['集合计划客户数'])
            ws.cell(row_num, 4, row['客户数合计'])

    # =========================================================================
    # KPI表写入
    # =========================================================================

    def write_kpi_scale(self, data: pd.DataFrame):
        """写入kpi-规模表"""
        ws = self.workbook['kpi-规模']

        for idx, row in data.iterrows():
            row_num = idx + 3  # 从第3行开始
            ws.cell(row_num, 1, row['机构名称'])

            # 当月数据
            ws.cell(row_num, 14, row.get('curr_trustee', 0))   # N列
            ws.cell(row_num, 15, row.get('curr_account', 0))   # O列
            ws.cell(row_num, 16, row.get('curr_invest', 0))    # P列
            # Q列(托管规模)保持手工填写

    def write_kpi_customer(self, data: pd.DataFrame):
        """写入kpi-客户数表"""
        ws = self.workbook['kpi-客户数']

        for idx, row in data.iterrows():
            row_num = idx + 3
            ws.cell(row_num, 1, row['机构名称'])
            ws.cell(row_num, 8, row.get('curr安心健养', 0))  # H列
            # I列(托管客户数)保持手工填写

    # =========================================================================
    # 基础数据表写入
    # =========================================================================

    def write_base_data(self, data: pd.DataFrame):
        """写入基础数据-养老金融客户、规模表"""
        ws = self.workbook['基础数据-养老金融客户、规模']

        for idx, row in data.iterrows():
            row_num = idx + 5  # 从第5行开始
            ws.cell(row_num, 1, row['机构名称'])

            # 当月数据 (AB-AN列, 即27-40列)
            base_col = 27
            for indicator, value in row.items():
                if indicator != '机构名称' and indicator.startswith('curr'):
                    ws.cell(row_num, base_col, value)
                    base_col += 1

    # =========================================================================
    # 辅助计算表写入
    # =========================================================================

    def write_aux_calc(self, data: pd.DataFrame):
        """写入辅助计算表"""
        ws = self.workbook['辅助计算']

        for idx, row in data.iterrows():
            row_num = idx + 5  # 从第5行开始
            ws.cell(row_num, 1, row['机构名称'])

            # B列: 当月客户数
            ws.cell(row_num, 2, row.get('curr_customers', 0))
            # E列: 当月规模
            ws.cell(row_num, 5, row.get('curr_scale', 0))
            # 排名列
            ws.cell(row_num, 12, row.get('customer_rank', 0))  # L列
            ws.cell(row_num, 15, row.get('scale_rank', 0))     # O列

    # =========================================================================
    # 通报表写入
    # =========================================================================

    def write_report_customer(self, data: pd.DataFrame):
        """写入通报-养老金融客户表"""
        ws = self.workbook['通报-养老金融客户']

        # 清空现有数据
        for row in range(4, 41):
            for col in range(1, 10):
                ws.cell(row, col, None)

        # 写入新数据
        for idx, row_data in enumerate(data.itertuples()):
            row_num = idx + 4
            # 排名(固定1-37)
            ws.cell(row_num, 1, idx + 1)
            ws.cell(row_num, 4, idx + 1)
            ws.cell(row_num, 7, idx + 1)

    def write_report_scale(self, data: pd.DataFrame):
        """写入通报-养老金融规模表"""
        ws = self.workbook['通报-养老金融规模']

        # 清空现有数据
        for row in range(4, 41):
            for col in range(1, 10):
                ws.cell(row, col, None)

        # 写入新数据
        for idx, row_data in enumerate(data.itertuples()):
            row_num = idx + 4
            ws.cell(row_num, 1, idx + 1)
            ws.cell(row_num, 4, idx + 1)
            ws.cell(row_num, 7, idx + 1)

    # =========================================================================
    # 读取手工数据
    # =========================================================================

    def read_manual_data(self) -> Dict:
        """读取所有手工维护的数据"""
        data = {
            'kpi_scale': {},
            'kpi_customer': {},
            'base_data': {}
        }

        # 读取kpi-规模表的手工列
        ws = self.workbook['kpi-规模']
        for row in range(3, 40):
            branch = ws.cell(row, 1).value
            if branch in BRANCHES:
                data['kpi_scale'][branch] = {
                    'curr托管': ws.cell(row, 17).value or 0,  # Q列
                    'last托管': ws.cell(row, 11).value or 0,  # K列
                    'nov托管': ws.cell(row, 5).value or 0,    # E列
                }

        # 读取kpi-客户数表的手工列
        ws = self.workbook['kpi-客户数']
        for row in range(3, 40):
            branch = ws.cell(row, 1).value
            if branch in BRANCHES:
                data['kpi_customer'][branch] = {
                    'curr托管': ws.cell(row, 9).value or 0,  # I列
                }

        return data

    # =========================================================================
    # 验证数据
    # =========================================================================

    def validate_data(self) -> Dict:
        """数据验证"""
        validation = {
            'total_branches': 0,
            'with_data': 0,
            'coverage': 0,
            'issues': []
        }

        # 检查基础数据表
        ws = self.workbook['基础数据-养老金融客户、规模']
        for row in range(5, 42):
            branch = ws.cell(row, 1).value
            if branch in BRANCHES:
                validation['total_branches'] += 1
                # 检查是否有数据
                has_data = False
                for col in range(27, 41):  # AB-AN列
                    if ws.cell(row, col).value and ws.cell(row, col).value != 0:
                        has_data = True
                        break
                if has_data:
                    validation['with_data'] += 1

        if validation['total_branches'] > 0:
            validation['coverage'] = validation['with_data'] / validation['total_branches']

        return validation


from typing import Tuple
