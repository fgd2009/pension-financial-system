# -*- coding: utf-8 -*-
"""
养老金融通报数据自动化处理系统
数据初始化模块

用于初始化模板中的分行列表和表头结构

作者: Matrix Agent
创建日期: 2026-03-19
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.config import BRANCHES, BRANCH_COUNT


def initialize_template(template_path: str, output_path: str = None) -> bool:
    """
    初始化模板
    - 填充37个分行列表
    - 设置公式
    - 格式化表头

    Args:
        template_path: 源模板路径
        output_path: 输出路径(如果为None,则覆盖原模板)

    Returns:
        bool: 是否成功
    """
    if output_path is None:
        output_path = template_path

    try:
        wb = load_workbook(template_path)

        # 初始化透视表
        _init_pivot_tables(wb)

        # 初始化KPI表
        _init_kpi_tables(wb)

        # 初始化基础数据表
        _init_base_data_table(wb)

        # 初始化辅助计算表
        _init_aux_calc_table(wb)

        # 初始化通报表
        _init_report_tables(wb)

        # 保存
        wb.save(output_path)
        print(f"模板初始化完成: {output_path}")
        return True

    except Exception as e:
        print(f"初始化失败: {e}")
        return False


def _init_pivot_tables(wb: openpyxl.Workbook):
    """初始化透视表"""
    # 透视-规模汇总
    ws = wb['透视-规模汇总']
    for idx, branch in enumerate(BRANCHES, 2):  # 从第2行开始
        ws.cell(idx, 1, branch)

    # 透视-客户数汇总
    ws = wb['透视-客户数汇总']
    for idx, branch in enumerate(BRANCHES, 2):
        ws.cell(idx, 1, branch)


def _init_kpi_tables(wb: openpyxl.Workbook):
    """初始化KPI表"""
    # kpi-规模
    ws = wb['kpi-规模']
    for idx, branch in enumerate(BRANCHES, 3):  # 从第3行开始
        ws.cell(idx, 1, branch)

    # kpi-客户数
    ws = wb['kpi-客户数']
    for idx, branch in enumerate(BRANCHES, 3):
        ws.cell(idx, 1, branch)


def _init_base_data_table(wb: openpyxl.Workbook):
    """初始化基础数据表"""
    ws = wb['基础数据-养老金融客户、规模']
    for idx, branch in enumerate(BRANCHES, 5):  # 从第5行开始
        ws.cell(idx, 1, branch)


def _init_aux_calc_table(wb: openpyxl.Workbook):
    """初始化辅助计算表"""
    ws = wb['辅助计算']
    for idx, branch in enumerate(BRANCHES, 5):  # 从第5行开始
        ws.cell(idx, 1, branch)


def _init_report_tables(wb: openpyxl.Workbook):
    """初始化通报表"""
    # 通报-养老金融客户
    ws = wb['通报-养老金融客户']
    for idx in range(4, 41):  # 从第4行开始
        ws.cell(idx, 1, idx - 3)  # 排名1-37
        ws.cell(idx, 4, idx - 3)
        ws.cell(idx, 7, idx - 3)

    # 通报-养老金融规模
    ws = wb['通报-养老金融规模']
    for idx in range(4, 41):
        ws.cell(idx, 1, idx - 3)
        ws.cell(idx, 4, idx - 3)
        ws.cell(idx, 7, idx - 3)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python init_template.py <模板路径> [输出路径]")
        sys.exit(1)

    template = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else None

    initialize_template(template, output)
